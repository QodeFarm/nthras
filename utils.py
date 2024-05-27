import csv
import json
from django.apps import apps
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from rest_framework import serializers
from rest_framework.serializers import Serializer

def get_model(model_name):
    """Get the model class based on model name."""
    try:
        app_label, model_name = model_name.split('.')
        return apps.get_model(app_label, model_name)
    except ValueError:
        raise ValueError("Model name must be in the format 'app_label.ModelName'")
    except LookupError:
        raise LookupError("Model not found. Check if the model name is correct and includes the app label.")


def get_serializer(model_name):
    """Create a serializer class for the given model."""
    class DynamicSerializer(serializers.ModelSerializer):
        class Meta:
            model = model_name
            fields = '__all__'
    
    return DynamicSerializer


def list_all_models():
    """List all registered models in the project."""
    models = []
    for app_config in apps.get_app_configs():
        for model in app_config.get_models():
            models.append(f"{app_config.label}.{model.__name__}")
    return models

def create_download_response(serializer: Serializer, filename: str = "data.json") -> HttpResponse:
    # Convert the serialized data to a formatted JSON string
    formatted_data = json.dumps(serializer.data, indent=4)

def create_excel_response(data, filename='data.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
 
    if data:
        # Write the headers
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
 
        # Write the data rows
        for row_num, item in enumerate(data, 2):
            for col_num, (key, value) in enumerate(item.items(), 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{row_num}'] = str(value)  # Convert value to string
 
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response